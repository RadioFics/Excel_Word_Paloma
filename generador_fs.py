"""
generador_fs.py — Traslada el Estado de Situación Financiera desde un libro
de Excel a la plantilla de Word, identificando la hoja y las columnas por su
CONTENIDO en vez de por posiciones fijas.

Qué cambió respecto de la versión rígida
----------------------------------------
1. La hoja ya NO tiene que llamarse exactamente "FS". Se usa ese nombre si
   existe; si no, se elige la hoja cuyo contenido tiene más marcadores de un
   estado de situación financiera ("Total assets", "ASSETS", "Situación
   Financiera", …). Configurable en config.json -> "hoja" / "hoja_marcadores".

2. Las columnas (etiqueta, nota, cifra actual, cifra comparativa, Tipo) se
   detectan por su contenido:
     - las dos columnas con más celdas numéricas -> cifras actual y previa;
     - la primera columna con mucho texto a su izquierda -> etiqueta;
     - una columna intermedia, casi vacía y con enteros pequeños, o cuyo
       encabezado dice "Note"/"Nota" -> nota;
     - una columna cuyo encabezado dice "Tipo"/"Type" -> Tipo (opcional).
   Se pueden forzar en config.json -> "columnas" (letras A, C, E, F, G).

3. La columna "Tipo" pasa a ser OPCIONAL. Si una fila no la trae, el
   programa INFIERE el tipo por señales de la propia hoja (negrita en las
   cifras, etiqueta sin cifras, texto "Total …", celda de texto largo sin
   cifras, etc.), genera el documento igual, y deja en
   salidas/revisar_tipos.csv el detalle fila por fila (tipo, origen y la
   señal que lo decidió) para que una persona lo revise.
   Ya NO se aborta solo porque falte la columna "Tipo".

Tipos de fila
-------------
  H = encabezado de sección           I = línea de detalle
  S = subtotal (sin etiqueta)         T = total (etiqueta "Total …")
  N = nota de texto libre, sin cifras  X = excluir a propósito

ADVERTENCIA OPERATIVA — edición del Excel
----------------------------------------
Abra y guarde el libro SIEMPRE desde Excel o LibreOffice. NUNCA reguarde
este libro con un script de openpyxl: openpyxl no recalcula fórmulas y, al
reguardar, descarta el valor cacheado de TODAS las fórmulas. El síntoma es
que el generador corre sin error pero el Word sale con las cifras en blanco.

Uso
---
    python generador_fs.py <libro.xlsx> [plantilla.docx]
    python generador_fs.py                 (busca el Excel por convención)
    python generador_fs.py <libro.xlsx> --revisar
                                           (solo escribe revisar_tipos.csv,
                                            no genera el Word)
Doble clic (Windows): use generar.bat.
"""
import sys
import re
import json
import csv
import hashlib
import traceback
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from docxtpl import DocxTemplate
from jinja2.sandbox import SandboxedEnvironment

# Rutas — funcionan tanto como script suelto como empaquetado en un .exe
# (PyInstaller --onefile). Cuando es .exe:
#   BASE     = carpeta donde está el .exe (ahí busca el Excel y escribe salidas/)
#   RECURSOS = carpeta temporal con los archivos embebidos (plantilla, config)
if getattr(sys, "frozen", False):
    BASE = Path(sys.executable).resolve().parent
    RECURSOS = Path(getattr(sys, "_MEIPASS", BASE))
else:
    BASE = Path(__file__).resolve().parent
    RECURSOS = BASE

SALIDAS = BASE / "salidas"
CONFIG_PATH = BASE / "config.json"            # config externa opcional (junto al .exe)
CONFIG_EMBEBIDA = RECURSOS / "config.json"    # config por defecto embebida
PLANTILLA_EMBEBIDA = RECURSOS / "plantilla_estado_situacion_financiera.docx"

TIPOS_VALIDOS = {"H", "I", "S", "T", "N"}

# Valores por defecto. config.json (si existe) los sobreescribe clave por clave.
DEFAULTS = {
    "empresa": "Collective Mining Ltd.",
    "hoja": "FS",
    "hoja_marcadores": [
        "situación financiera", "situacion financiera",
        "statement of financial position", "financial position",
        "total assets", "total liabilities and equity",
        "assets", "liabilities and equity",
    ],
    "plantilla": "plantilla_estado_situacion_financiera.docx",
    "buscar_por_convencion": "FS",
    "primera_fila": "auto",          # un entero fija la fila de inicio; "auto" la detecta
    "columnas": {                     # letras (A, C, E, F, G) para forzar; null = detectar
        "etiqueta": None, "nota": None, "actual": None, "previo": None, "tipo": None,
    },
    "marcadores_excluir": ["control check", "check", "cuadre", "balance check"],
    "max_filas_scan": 400,
    "max_cols_scan": 16,
}


# --------------------------------------------------------------------------- #
#  Formato de valores (sin cambios respecto de la versión anterior)
# --------------------------------------------------------------------------- #
def num(valor):
    """Formato contable: negativos entre paréntesis, sin decimales.
    Los marcadores de cero ('-', 'ꟷ') se devuelven tal cual."""
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor
    if not isinstance(valor, (int, float)):
        return str(valor)
    if valor < 0:
        return f"({abs(valor):,.0f})"
    return f"{valor:,.0f}"


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip("() ")


def sanear(texto_):
    """Impide que un valor del Excel se convierta en una ruta peligrosa."""
    return re.sub(r"[^A-Za-z0-9_\-]", "_", str(texto_))[:40]


# --------------------------------------------------------------------------- #
#  Utilidades de detección
# --------------------------------------------------------------------------- #
def _norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


def _es_numero(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _texto_no_vacio(v):
    return isinstance(v, str) and v.strip() != ""


def _marcador_cero(v):
    return isinstance(v, str) and v.strip() in {"-", "–", "—", "−", "ꟷ", "‑", "·", "0"}


def _col_por_letra(x):
    if x is None or str(x).strip() == "":
        return None
    return column_index_from_string(str(x).strip().upper())


def _primer_texto(valores, col, filas):
    if not col:
        return ""
    for r in filas:
        if 0 < r < len(valores) and _texto_no_vacio(valores[r][col]):
            return valores[r][col]
    return ""


def _fusionar_config(cfg, ruta):
    if not ruta.exists():
        return
    try:
        usuario = json.loads(ruta.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise ValueError(f"{ruta.name} tiene un error de sintaxis:\n  {e}")
    for k, v in usuario.items():
        if k.startswith("_"):
            continue
        if k == "columnas" and isinstance(v, dict):
            cfg["columnas"].update(v)
        else:
            cfg[k] = v


def cargar_config():
    cfg = json.loads(json.dumps(DEFAULTS))  # copia profunda
    _fusionar_config(cfg, CONFIG_EMBEBIDA)          # 1) config embebida en el .exe
    if CONFIG_PATH != CONFIG_EMBEBIDA:
        _fusionar_config(cfg, CONFIG_PATH)          # 2) config externa junto al .exe (gana)
    return cfg


# --------------------------------------------------------------------------- #
#  Elección de la hoja
# --------------------------------------------------------------------------- #
def _elegir_hoja_por_contenido(wb, cfg):
    marcadores = [_norm(m) for m in cfg["hoja_marcadores"] if _norm(m)]
    convencion = _norm(cfg.get("buscar_por_convencion"))
    mejor_nombre, mejor_score = None, 0

    for nombre in wb.sheetnames:
        ws = wb[nombre]
        try:
            texto_hoja = " ".join(
                _norm(v)
                for fila in ws.iter_rows(max_row=60, max_col=8, values_only=True)
                for v in fila
                if _texto_no_vacio(v)
            )
        except Exception:
            continue
        score = sum(1 for m in marcadores if m in texto_hoja)
        if convencion and convencion in _norm(nombre):
            score += 1
        try:
            if getattr(ws, "sheet_state", "visible") == "visible":
                score += 0.5
        except Exception:
            pass
        if score > mejor_score:
            mejor_nombre, mejor_score = nombre, score

    if mejor_nombre is not None and mejor_score >= 2:
        return mejor_nombre, f"'{mejor_nombre}' elegida por contenido ({mejor_score:g} señales)"

    raise ValueError(
        "No pude identificar la hoja del Estado de Situación Financiera.\n"
        f"No existe una hoja llamada '{cfg['hoja']}' y ninguna otra hoja tiene\n"
        "suficientes marcadores de contenido (p. ej. 'Total assets', 'ASSETS',\n"
        "'Situación Financiera').\n"
        "Indique el nombre exacto en config.json -> \"hoja\"."
    )


# --------------------------------------------------------------------------- #
#  Materialización de la hoja elegida (una sola pasada)
# --------------------------------------------------------------------------- #
def _materializar(ws, cfg):
    max_filas, max_cols = cfg["max_filas_scan"], cfg["max_cols_scan"]
    valores = [[None] * (max_cols + 1)]
    negrita = [[False] * (max_cols + 1)]
    for fila in ws.iter_rows(min_row=1, max_row=max_filas, max_col=max_cols):
        vrow = [None] * (max_cols + 1)
        brow = [False] * (max_cols + 1)
        for c, celda in enumerate(fila, start=1):
            if c > max_cols:
                break
            vrow[c] = celda.value
            try:
                brow[c] = bool(celda.font and celda.font.bold)
            except Exception:
                brow[c] = False
        valores.append(vrow)
        negrita.append(brow)

    n_filas = len(valores) - 1
    while n_filas > 1 and all(v in (None, "") for v in valores[n_filas][1:]):
        n_filas -= 1
    return valores, negrita, n_filas


# --------------------------------------------------------------------------- #
#  Detección de columnas por contenido
# --------------------------------------------------------------------------- #
def detectar_columnas(valores, cfg, n_filas, max_cols):
    forzado = {k: _col_por_letra(v) for k, v in cfg["columnas"].items()}
    col = dict(forzado)

    # 1. encabezados explícitos en las primeras filas
    for r in range(1, min(8, n_filas) + 1):
        for c in range(1, max_cols + 1):
            t = _norm(valores[r][c])
            if not t:
                continue
            if col.get("tipo") is None and t in ("tipo", "type"):
                col["tipo"] = c
            if col.get("nota") is None and t in ("note", "nota", "notes", "notas"):
                col["nota"] = c

    # 2. perfil de cada columna sobre el cuerpo de la hoja
    perfil = {}
    for c in range(1, max_cols + 1):
        n_txt = n_num = 0
        for r in range(2, n_filas + 1):
            v = valores[r][c]
            if _es_numero(v):
                n_num += 1
            elif _texto_no_vacio(v):
                n_txt += 1
        perfil[c] = (n_txt, n_num)

    # 3. columnas de cifras: las dos con más numéricos (empate -> más a la derecha)
    if col.get("actual") is None or col.get("previo") is None:
        con_numeros = [c for c in perfil if perfil[c][1] >= 2]
        con_numeros.sort(key=lambda c: (perfil[c][1], c), reverse=True)
        elegidas = sorted(con_numeros[:2])
        if len(elegidas) >= 2:
            col["actual"] = col.get("actual") or elegidas[0]
            col["previo"] = col.get("previo") or elegidas[1]
        elif len(elegidas) == 1:
            col["actual"] = col.get("actual") or elegidas[0]

    # 4. etiqueta: primera columna con bastante texto, a la izquierda de las cifras
    if col.get("etiqueta") is None:
        limite = col.get("actual") or (max_cols + 1)
        textuales = [c for c in range(1, limite) if perfil[c][0] >= 3]
        col["etiqueta"] = textuales[0] if textuales else 1

    # 5. nota: columna intermedia casi vacía con enteros pequeños
    if col.get("nota") is None and col.get("actual"):
        for c in range(col["etiqueta"] + 1, col["actual"]):
            n_txt, _ = perfil[c]
            enteros_peq = sum(
                1 for r in range(2, n_filas + 1)
                if _es_numero(valores[r][c]) and float(valores[r][c]).is_integer()
                and 0 < valores[r][c] <= 99
            )
            if enteros_peq >= 1 and n_txt <= 2:
                col["nota"] = c
                break

    if not col.get("actual"):
        raise ValueError(
            "No pude identificar las columnas de cifras en la hoja.\n"
            "Defina las letras a mano en config.json -> \"columnas\"\n"
            "(etiqueta / nota / actual / previo / tipo)."
        )

    return {
        "etiqueta": col.get("etiqueta"),
        "nota": col.get("nota"),
        "actual": col.get("actual"),
        "previo": col.get("previo"),
        "tipo": col.get("tipo"),
    }


# --------------------------------------------------------------------------- #
#  Región de datos (fila inicial y final)
# --------------------------------------------------------------------------- #
def detectar_region(valores, cols, cfg, n_filas):
    ce, ca, cp = cols["etiqueta"], cols["actual"], cols["previo"]
    pf = cfg["primera_fila"]

    if isinstance(pf, int) and pf > 0:
        primera = pf
    else:
        primera = None
        for r in range(1, n_filas + 1):
            et = valores[r][ce]
            va = valores[r][ca] if ca else None
            vp = valores[r][cp] if cp else None
            if not _texto_no_vacio(et):
                continue
            if _norm(et) in ("", "$"):
                continue
            if r > 4 or _es_numero(va) or _es_numero(vp):
                primera = r
                break
        if primera is None:
            primera = 5

    ultima = primera
    for r in range(primera, n_filas + 1):
        if (_texto_no_vacio(valores[r][ce])
                or (ca and _es_numero(valores[r][ca]))
                or (cp and _es_numero(valores[r][cp]))):
            ultima = r
    return primera, ultima


# --------------------------------------------------------------------------- #
#  Encabezado (título, fechas, unidad, moneda)
# --------------------------------------------------------------------------- #
def leer_encabezado(valores, cols, cfg, primera):
    ce, ca, cp = cols["etiqueta"], cols["actual"], cols["previo"]
    filas_hdr = list(range(1, max(primera, 2)))

    def celda(c, offset):
        r = 1 + offset
        if c and 0 < r < primera and r < len(valores):
            return valores[r][c]
        return None

    return {
        "empresa": cfg["empresa"],
        "titulo": _primer_texto(valores, ce, filas_hdr) or "",
        "fecha_actual": celda(ca, 0) or "",
        "fecha_previa": celda(cp, 0) or "",
        "miles": celda(ca, 1) or "",
        "estado_actual": texto(celda(ca, 2)),
        "estado_previo": texto(celda(cp, 2)),
        "moneda": celda(ca, 3) or "",
    }


# --------------------------------------------------------------------------- #
#  Inferencia del tipo de fila
# --------------------------------------------------------------------------- #
def inferir_tipo(vals, bold, cols, cfg):
    ce, cn, ca, cp = cols["etiqueta"], cols["nota"], cols["actual"], cols["previo"]
    et = vals[ce]
    nota = vals[cn] if cn else None
    va = vals[ca] if ca else None
    vp = vals[cp] if cp else None

    et_txt = et.strip() if isinstance(et, str) else ("" if et is None else str(et))
    et_norm = _norm(et_txt)
    nota_norm = _norm(nota)
    hay_valor = (_es_numero(va) or _es_numero(vp)
                 or _marcador_cero(va) or _marcador_cero(vp))
    tiene_nota = _es_numero(nota) or _texto_no_vacio(str(nota) if nota is not None else "")
    en_negrita = bool((ca and bold[ca]) or (cp and bold[cp]) or (ce and bold[ce]))

    # exclusión: el marcador puede estar en la etiqueta o en la columna de nota
    for m in cfg["marcadores_excluir"]:
        mm = _norm(m)
        if mm and (mm in et_norm or mm in nota_norm):
            return "X", "fila de control/cuadre"

    if not et_txt and not hay_valor and not tiene_nota:
        return None, "fila vacía"

    if et_norm.startswith("total"):
        return "T", "empieza por 'Total'"

    if not et_txt and hay_valor:
        return "S", "cifras sin etiqueta (subtotal)"

    if et_txt and not hay_valor:
        if et_txt.isupper() or et_txt.endswith(":") or en_negrita or len(et_txt) <= 40:
            return "H", "etiqueta sin cifras (encabezado)"
        return "N", "texto largo sin cifras (nota)"

    if et_txt and (hay_valor or tiene_nota):
        return "I", "etiqueta con cifras o nota"

    return None, "sin señales suficientes"


# --------------------------------------------------------------------------- #
#  Construcción de las líneas + registro de revisión
# --------------------------------------------------------------------------- #
def construir_lineas(valores, negrita, cols, cfg, primera, ultima, hay_col_tipo):
    ce, cn, ca, cp, ct = (cols["etiqueta"], cols["nota"], cols["actual"],
                          cols["previo"], cols["tipo"])
    lineas, revision = [], []
    sin_tipo, tipo_invalido = [], []
    n_declarados = n_inferidos = 0

    for r in range(primera, ultima + 1):
        vals, bold = valores[r], negrita[r]
        crudo = vals[ct] if ct else None
        declarado = str(crudo).strip().upper() if crudo not in (None, "") else ""

        et_prev = vals[ce] if _texto_no_vacio(vals[ce]) else ""

        if declarado == "X":
            revision.append((r, et_prev, "", "", "", "X", "declarado", "excluido a propósito"))
            continue

        if declarado in TIPOS_VALIDOS:
            tipo, origen, senal = declarado, "declarado", ""
        elif declarado:
            tipo_invalido.append((r, crudo))
            inf, senal = inferir_tipo(vals, bold, cols, cfg)
            if inf in (None, "X"):
                revision.append((r, et_prev, "", "", "", declarado, "tipo inválido",
                                 f"'{crudo}' no es H/I/S/T/N; {senal}"))
                continue
            tipo, origen = inf, "inferido (Tipo inválido)"
        else:
            inf, senal = inferir_tipo(vals, bold, cols, cfg)
            if inf is None:
                continue
            if inf == "X":
                revision.append((r, et_prev, "", "", "", "", "inferido", senal))
                continue
            tipo, origen = inf, "inferido"
            if hay_col_tipo:
                sin_tipo.append(r)

        etiqueta = vals[ce] or ""
        nota = texto(vals[cn]) if cn else ""
        actual = num(vals[ca]) if ca else ""
        previo = num(vals[cp]) if cp else ""
        lineas.append({"tipo": tipo, "etiqueta": etiqueta, "nota": nota,
                       "actual": actual, "previo": previo})
        revision.append((r, etiqueta, nota, actual, previo, tipo, origen, senal))
        if origen == "declarado":
            n_declarados += 1
        else:
            n_inferidos += 1

    return lineas, revision, sin_tipo, tipo_invalido, n_declarados, n_inferidos


def escribir_revision(revision):
    SALIDAS.mkdir(exist_ok=True)
    ruta = SALIDAS / "revisar_tipos.csv"
    with ruta.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["fila", "etiqueta", "nota", "actual", "previo",
                    "tipo", "origen", "señal"])
        for row in revision:
            w.writerow(row)
    return ruta


# --------------------------------------------------------------------------- #
#  Orquestador de lectura
# --------------------------------------------------------------------------- #
def leer_contexto(ruta_xlsx, cfg):
    wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
    try:
        objetivo = cfg["hoja"]
        if objetivo and objetivo in wb.sheetnames:
            ws = wb[objetivo]
            como_hoja = f"'{objetivo}' por nombre exacto"
        else:
            nombre, como_hoja = _elegir_hoja_por_contenido(wb, cfg)
            wb.close()
            wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
            ws = wb[nombre]

        hoja_titulo = ws.title
        valores, negrita, n_filas = _materializar(ws, cfg)
        cols = detectar_columnas(valores, cfg, n_filas, cfg["max_cols_scan"])
        primera, ultima = detectar_region(valores, cols, cfg, n_filas)
        ctx = leer_encabezado(valores, cols, cfg, primera)
        hay_col_tipo = cols["tipo"] is not None
        (lineas, revision, sin_tipo, tipo_invalido,
         n_declarados, n_inferidos) = construir_lineas(
            valores, negrita, cols, cfg, primera, ultima, hay_col_tipo)
    finally:
        wb.close()

    if not lineas:
        raise ValueError(
            f"Identifiqué la hoja ({hoja_titulo}, {como_hoja}) y las columnas,\n"
            f"pero no obtuve ninguna línea en la región filas {primera}–{ultima}.\n"
            "Posibles causas: el Excel se guardó sin valores cacheados de las\n"
            "fórmulas, o las columnas detectadas no son las correctas.\n"
            "Revise config.json (columnas / primera_fila)."
        )

    ctx["lineas"] = lineas
    ctx["_meta"] = {
        "hoja": hoja_titulo,
        "como_hoja": como_hoja,
        "columnas": {k: (get_column_letter(v) if v else "—") for k, v in cols.items()},
        "region": (primera, ultima),
        "hay_col_tipo": hay_col_tipo,
        "n_declarados": n_declarados,
        "n_inferidos": n_inferidos,
    }
    ctx["_avisos"] = {
        "sin_tipo": sin_tipo,
        "tipo_invalido": tipo_invalido,
        "revision": revision,
    }
    return ctx


def encontrar_excel_por_convencion(cfg):
    clave = cfg["buscar_por_convencion"]
    candidatos = sorted(
        [p for p in BASE.glob("*.xlsx") if clave.lower() in p.stem.lower()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidatos:
        icono = "GeneradorFS.exe" if getattr(sys, "frozen", False) else "generar.bat"
        raise ValueError(
            "No encontré ningún Excel en esta carpeta.\n\n"
            f"Haga una de estas dos cosas:\n"
            f"  1) Arrastre su archivo de Excel encima de {icono}, o\n"
            f"  2) Copie el Excel a esta misma carpeta (su nombre debe contener "
            f"'{clave}') y vuelva a ejecutar."
        )
    return candidatos[0]


# --------------------------------------------------------------------------- #
#  Ejecución
# --------------------------------------------------------------------------- #
def ejecutar(argv):
    cfg = cargar_config()
    args = [a for a in argv[1:] if not a.startswith("--")]
    flags = {a.lower() for a in argv[1:] if a.startswith("--")}
    solo_revisar = "--revisar" in flags

    if args:
        xlsx = Path(args[0]).resolve()
    else:
        xlsx = encontrar_excel_por_convencion(cfg)
        print(f"(Sin archivo indicado: usando '{xlsx.name}' por convención de nombre)")

    if len(args) >= 2:
        plantilla = Path(args[1]).resolve()
    else:
        plantilla = BASE / cfg["plantilla"]
        if not plantilla.exists():
            plantilla = PLANTILLA_EMBEBIDA          # la que viaja dentro del .exe

    if not xlsx.exists():
        raise ValueError(f"No se encontró el libro de Excel:\n  {xlsx}")
    if not solo_revisar and not plantilla.exists():
        raise ValueError(f"No se encontró la plantilla de Word:\n  {plantilla}")

    ctx = leer_contexto(xlsx, cfg)
    meta = ctx.pop("_meta")
    avisos = ctx.pop("_avisos")

    sello = datetime.now().strftime("%Y%m%d-%H%M")
    csv_ruta = escribir_revision(avisos["revision"])

    destino = None
    if not solo_revisar:
        entorno = SandboxedEnvironment(autoescape=True)
        doc = DocxTemplate(plantilla)
        doc.render(ctx, entorno)
        sha = hashlib.sha256(xlsx.read_bytes()).hexdigest()[:12]
        doc.docx.core_properties.comments = (
            f"origen={xlsx.name} sha256={sha} hoja={meta['hoja']} "
            f"plantilla={plantilla.name} generado={sello}"
        )
        SALIDAS.mkdir(exist_ok=True)
        destino = SALIDAS / f"estado_situacion_financiera_{sanear(ctx['fecha_actual'])}_{sello}.docx"
        doc.save(destino)

    cols = meta["columnas"]
    print()
    print("=" * 64)
    print(" REVISIÓN GENERADA" if solo_revisar else " DOCUMENTO GENERADO CORRECTAMENTE")
    print("=" * 64)
    if destino:
        print(f" Archivo:            {destino.name}")
        print(f" Carpeta:            {destino.parent}")
    print(f" Hoja usada:         {meta['hoja']}  ({meta['como_hoja']})")
    print(f" Columnas:           etiqueta={cols['etiqueta']}  nota={cols['nota']}  "
          f"actual={cols['actual']}  previo={cols['previo']}  tipo={cols['tipo']}")
    print(f" Región de datos:    filas {meta['region'][0]}–{meta['region'][1]}")
    print(f" Líneas trasladadas: {len(ctx['lineas'])}   "
          f"({meta['n_declarados']} declaradas, {meta['n_inferidos']} inferidas)")
    print(f" Revisión de tipos:  {csv_ruta}")

    if not meta["hay_col_tipo"]:
        print()
        print(" NOTA — la hoja no tiene columna 'Tipo': TODOS los tipos se")
        print("        infirieron. Abra revisar_tipos.csv y confírmelos antes")
        print("        de usar el documento. Para fijarlos, agregue una columna")
        print("        'Tipo' a la hoja con H/I/S/T/N/X.")

    if avisos["sin_tipo"]:
        print()
        print(" AVISO — filas con datos pero SIN 'Tipo' declarado (se infirió):")
        print("   Fila(s):", ", ".join(str(r) for r in avisos["sin_tipo"]))

    if avisos["tipo_invalido"]:
        print()
        print(" AVISO — filas con un 'Tipo' no reconocido (se infirió en su lugar):")
        for r, t in avisos["tipo_invalido"]:
            print(f"   Fila {r}: '{t}'  (válidos: H, I, S, T, N, X)")

    print("=" * 64)


def _pausa():
    """Espera Enter solo si hay una consola interactiva (doble clic / .bat).
    Si el programa se llama desde un script o una tubería, no pausa."""
    try:
        if sys.stdin and sys.stdin.isatty():
            input("\nPresione Enter para cerrar...")
    except (EOFError, OSError):
        pass


def main():
    try:
        ejecutar(sys.argv)
    except ValueError as e:
        print()
        print("=" * 64)
        print(" NO SE PUDO GENERAR EL DOCUMENTO")
        print("=" * 64)
        print(str(e))
        print("=" * 64)
        _pausa()
        sys.exit(1)
    except Exception:
        print()
        print("=" * 64)
        print(" OCURRIÓ UN ERROR INESPERADO — copie este texto para soporte")
        print("=" * 64)
        traceback.print_exc()
        print("=" * 64)
        _pausa()
        sys.exit(1)
    else:
        _pausa()


if __name__ == "__main__":
    main()
