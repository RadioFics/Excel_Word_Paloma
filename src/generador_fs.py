"""
generador_fs.py — Traslada el Estado de Situación Financiera desde un libro
de Excel a la plantilla de Word, identificando la hoja y las columnas por su
CONTENIDO en vez de por posiciones fijas.

Qué cambió respecto de la versión rígida
----------------------------------------
1. La hoja ya NO tiene que llamarse exactamente "FS". Se usa ese nombre si
   existe; si no, se elige la hoja cuyo contenido tiene más marcadores de un
   estado de situación financiera ("Total assets", "ASSETS", "Situación
   Financiera", …). Configurable en config.json -> "hoja" / "hoja_marcadores".

2. Las columnas (etiqueta, nota, cifra actual, cifra comparativa, Tipo) se
   detectan por su contenido:
     - las dos columnas con más celdas numéricas -> cifras actual y previa;
     - la primera columna con mucho texto a su izquierda -> etiqueta;
     - una columna intermedia, casi vacía y con enteros pequeños, o cuyo
       encabezado dice "Note"/"Nota" -> nota;
     - una columna cuyo encabezado dice "Tipo"/"Type" -> Tipo (opcional).
   Se pueden forzar en config.json -> "columnas" (letras A, C, E, F, G).

3. La columna "Tipo" pasa a ser OPCIONAL. Si una fila no la trae, el
   programa INFIERE el tipo por señales de la propia hoja (negrita en las
   cifras, etiqueta sin cifras, texto "Total …", celda de texto largo sin
   cifras, etc.), genera el documento igual, y deja en
   salidas/revisar_tipos.csv el detalle fila por fila (tipo, origen y la
   señal que lo decidió) para que una persona lo revise.
   Ya NO se aborta solo porque falte la columna "Tipo".

Tipos de fila
-------------
  H = encabezado de sección           I = línea de detalle
  S = subtotal (sin etiqueta)         T = total (etiqueta "Total …")
  N = nota de texto libre, sin cifras  X = excluir a propósito

ADVERTENCIA OPERATIVA — edición del Excel
----------------------------------------
Abra y guarde el libro SIEMPRE desde Excel o LibreOffice. NUNCA reguarde
este libro con un script de openpyxl: openpyxl no recalcula fórmulas y, al
reguardar, descarta el valor cacheado de TODAS las fórmulas. El síntoma es
que el generador corre sin error pero el Word sale con las cifras en blanco.

Uso
---
    python generador_fs.py <libro.xlsx> [plantilla.docx]
    python generador_fs.py                 (busca el Excel por convención)
    python generador_fs.py <libro.xlsx> --revisar
                                           (solo escribe revisar_tipos.csv,
                                            no genera el Word)
Doble clic (Windows): use generar.bat.
"""
import sys
import os
import re
import json
import csv
import hashlib
import traceback
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from docxtpl import DocxTemplate
from jinja2.sandbox import SandboxedEnvironment

# Rutas — funcionan tanto como script suelto como empaquetado en un .exe
# (PyInstaller --onefile). Cuando es .exe:
#   BASE     = carpeta donde está el .exe (ahí busca el Excel y escribe salidas/)
#   RECURSOS = carpeta temporal con los archivos embebidos (plantilla, config)
# Como script, el código vive en src\ pero BASE es la RAÍZ del proyecto: es
# ahí donde están config.json, salidas\ y las carpetas de recursos.
if getattr(sys, "frozen", False):
    BASE = Path(sys.executable).resolve().parent
    # El .exe suele quedarse donde lo dejó PyInstaller: dentro de dist\, un
    # nivel por debajo del proyecto. Si ahí no hay config.json pero sí lo
    # hay justo encima, la carpeta del proyecto es la de encima. Sin esto el
    # .exe lee la config EMBEBIDA (la del día que se compiló), ignora los
    # cambios de config.json y escribe salidas\ y la bitácora en dist\.
    if not (BASE / "config.json").exists() and (BASE.parent / "config.json").exists():
        BASE = BASE.parent
    RECURSOS = Path(getattr(sys, "_MEIPASS", BASE))
else:
    BASE = Path(__file__).resolve().parent.parent
    RECURSOS = BASE

SALIDAS = BASE / "salidas"
CONFIG_PATH = BASE / "config.json"            # config externa opcional (junto al .exe)
CONFIG_EMBEBIDA = RECURSOS / "config.json"    # config por defecto embebida

#: Ajustes de ESTE equipo. No se versiona (ver .gitignore) y gana sobre todo
#: lo demas. Existe porque config.json sí viaja por git: si ahí dentro hay
#: una ruta absoluta de una máquina, cada «git pull» la impone en la otra y
#: las dos se pisan para siempre. Todo lo que dependa del equipo va aquí.
CONFIG_LOCAL = BASE / "config.local.json"

#: Subcarpetas donde se buscan los recursos cuando el proyecto está
#: desplegado como código. Dentro del .exe todo queda plano en _MEIPASS,
#: por eso también se prueba la raíz.
_CARPETAS_RECURSO = ("", "plantillas", "ejemplos")


def buscar_recurso(nombre, *raices):
    """Localiza un archivo de apoyo (plantilla, ejemplo) sin importar si
    estamos como script (subcarpetas) o dentro del .exe (todo plano)."""
    for raiz in (raices or (BASE, RECURSOS)):
        for sub in _CARPETAS_RECURSO:
            cand = (raiz / sub / nombre) if sub else (raiz / nombre)
            if cand.exists():
                return cand
    return None

TIPOS_VALIDOS = {"H", "I", "S", "T", "N"}

# Valores por defecto. config.json (si existe) los sobreescribe clave por clave.
DEFAULTS = {
    "empresa": "Collective Mining Ltd.",
    "hoja": "FS",
    "hoja_marcadores": [
        "situación financiera", "situacion financiera",
        "statement of financial position", "financial position",
        "total assets", "total liabilities and equity",
        "assets", "liabilities and equity",
    ],
    "plantilla": "plantilla_estado_situacion_financiera.docx",
    "buscar_por_convencion": "FS",
    "primera_fila": "auto",          # un entero fija la fila de inicio; "auto" la detecta
    "columnas": {                     # letras (A, C, E, F, G) para forzar; null = detectar
        "etiqueta": None, "nota": None, "actual": None, "previo": None, "tipo": None,
    },
    "marcadores_excluir": ["control check", "check", "cuadre", "balance check"],
    "max_filas_scan": 400,
    "max_cols_scan": 16,
    # Rangos con nombre: un nombre de Excel "fs_total_assets" apunta a la
    # celda de etiqueta de esa fila y le da una identidad ESTABLE. Si alguien
    # renombra la fila o inserta filas encima, el nombre sigue apuntando a la
    # misma línea y el vínculo con el Word no se rompe. Ver CONTRATO.md.
    "prefijo_rangos": "fs_",
    # Documento de Word que se refresca en el sitio (el que vive en OneDrive).
    # Ruta absoluta, o relativa a esta carpeta. Si está vacío, hay que
    # indicar el documento a mano en cada orden. Solo lo usa fs_documento.py
    # / RefrescarFS.exe; el generador clásico lo ignora.
    "documento_base": "",
    # Dónde se anota el histórico de actualizaciones:
    #   "archivo"   -> un .log aparte (por defecto; deja el Word limpio)
    #   "documento" -> dentro del propio .docx, región fs-registro
    #   "ambos" / "no"
    "bitacora": "archivo",
    # Ruta del .log. Vacío = salidas\bitacora_<documento>.log
    "bitacora_archivo": "",
    # Aspecto de las regiones de datos en Word:
    #   "boundingBox" -> recuadro gris (se ve qué mantiene el Excel)
    #   "hidden"      -> la cifra se lee como texto normal del párrafo
    "apariencia_datos": "boundingBox",
    # Clave de la protección de documento que aplica la opción «proteger las
    # cifras». No es un secreto: solo evita ediciones accidentales.
    "clave_proteccion": "fs",
}


# --------------------------------------------------------------------------- #
#  Formato de valores (sin cambios respecto de la versión anterior)
# --------------------------------------------------------------------------- #
def num(valor):
    """Formato contable: negativos entre paréntesis, sin decimales.
    Los marcadores de cero ('-', 'ꟷ') se devuelven tal cual."""
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor
    if not isinstance(valor, (int, float)):
        return str(valor)
    if valor < 0:
        return f"({abs(valor):,.0f})"
    return f"{valor:,.0f}"


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip("() ")


def sanear(texto_):
    """Impide que un valor del Excel se convierta en una ruta peligrosa."""
    return re.sub(r"[^A-Za-z0-9_\-]", "_", str(texto_))[:40]


# --------------------------------------------------------------------------- #
#  Utilidades de detección
# --------------------------------------------------------------------------- #
def _norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


def _es_numero(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _texto_no_vacio(v):
    return isinstance(v, str) and v.strip() != ""


def _marcador_cero(v):
    return isinstance(v, str) and v.strip() in {"-", "–", "—", "−", "ꟷ", "‑", "·", "0"}


def _col_por_letra(x):
    if x is None or str(x).strip() == "":
        return None
    return column_index_from_string(str(x).strip().upper())


def _primer_texto(valores, col, filas):
    if not col:
        return ""
    for r in filas:
        if 0 < r < len(valores) and _texto_no_vacio(valores[r][col]):
            return valores[r][col]
    return ""


def _fusionar_config(cfg, ruta):
    if not ruta.exists():
        return
    try:
        usuario = json.loads(ruta.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise ValueError(f"{ruta.name} tiene un error de sintaxis:\n  {e}")
    for k, v in usuario.items():
        if k.startswith("_"):
            continue
        if k == "columnas" and isinstance(v, dict):
            cfg["columnas"].update(v)
        else:
            cfg[k] = v


def cargar_config():
    """Los ajustes, de menos a más prioritario.

    El orden importa: lo de este equipo (config.local.json) tiene que poder
    ganarle a lo que venga por git, o cada «pull» reimpondria las rutas de
    la otra máquina.
    """
    cfg = json.loads(json.dumps(DEFAULTS))  # copia profunda
    _fusionar_config(cfg, CONFIG_EMBEBIDA)          # 1) config embebida en el .exe
    if CONFIG_PATH != CONFIG_EMBEBIDA:
        _fusionar_config(cfg, CONFIG_PATH)          # 2) config del proyecto (viaja por git)
    if CONFIG_LOCAL != CONFIG_EMBEBIDA:
        _fusionar_config(cfg, CONFIG_LOCAL)         # 3) este equipo (no se versiona)
    return cfg


def raiz_onedrive():
    """La carpeta de OneDrive de empresa de este equipo, o None.

    No sirve la variable de entorno %OneDrive% a secas: en un equipo con las
    dos cuentas apunta a la PERSONAL ('C:\\Users\\x\\OneDrive'), no a la de
    la organizacion ('C:\\Users\\x\\OneDrive - Empresa'), que es donde vive
    el documento. Se prefiere siempre la de empresa.
    """
    casa = Path.home()
    empresa = sorted(p for p in casa.glob("OneDrive - *") if p.is_dir())
    if empresa:
        return empresa[0]
    for var in ("OneDriveCommercial", "OneDrive", "OneDriveConsumer"):
        v = os.environ.get(var)
        if v and Path(v).is_dir():
            return Path(v)
    suelta = casa / "OneDrive"
    return suelta if suelta.is_dir() else None


def expandir_ruta(crudo):
    """Convierte una ruta de config.json en algo válido en ESTE equipo.

    Entiende marcadores que no dependen de la máquina, para que la misma
    config sirva en todas:

        ${ONEDRIVE}\\Informes\\EF.docx   -> el OneDrive de empresa de aquí
        ${USUARIO}\\Documentos\\EF.docx  -> la carpeta del usuario actual
        ${PROYECTO}\\plantillas\\x.docx  -> la carpeta del proyecto/.exe
        ~\\Documentos\\EF.docx           -> igual que ${USUARIO}

    Una ruta normal se devuelve tal cual: nada de esto es obligatorio.
    """
    s = str(crudo or "").strip()
    if not s:
        return s
    onedrive = raiz_onedrive()
    reemplazos = {
        "${ONEDRIVE}": str(onedrive) if onedrive else str(Path.home()),
        "${USUARIO}": str(Path.home()),
        "${USERPROFILE}": str(Path.home()),
        "${PROYECTO}": str(BASE),
    }
    for marca, valor in reemplazos.items():
        if marca.lower() in s.lower():
            # sin distinguir mayusculas: nadie deberia fallar por escribir
            # ${onedrive} en minuscula
            i = s.lower().index(marca.lower())
            s = s[:i] + valor + s[i + len(marca):]
    if s.startswith("~"):
        s = str(Path.home()) + s[1:]
    return s


def compactar_ruta(ruta):
    """La inversa: sustituye la parte dependiente del equipo por su marcador.

    Se usa al GUARDAR el documento elegido, para que lo que quede escrito
    valga también en la otra máquina.
    """
    ruta = Path(ruta)
    onedrive = raiz_onedrive()
    for marca, raiz in (("${ONEDRIVE}", onedrive), ("${USUARIO}", Path.home())):
        if raiz is None:
            continue
        try:
            return marca + "\\" + str(ruta.relative_to(raiz))
        except ValueError:
            continue
    return str(ruta)


# --------------------------------------------------------------------------- #
#  Elección de la hoja
# --------------------------------------------------------------------------- #
#: Cuántas señales de contenido bastan para dar una hoja por buena.
UMBRAL_HOJA = 2


def _texto_de_hoja(ws, max_filas=120, max_cols=60):
    """El texto de la hoja, en una sola cadena normalizada.

    La ventana es ancha a propósito: antes se miraban 60 filas por 8
    columnas, y una tabla movida a K14 caía entera fuera. La hoja se
    puntuaba con cero señales y se descartaba, aunque fuese la buena.
    """
    try:
        return " ".join(
            _norm(v)
            for fila in ws.iter_rows(max_row=max_filas, max_col=max_cols,
                                     values_only=True)
            for v in fila
            if _texto_no_vacio(v)
        )
    except Exception:
        return ""


def puntuar_hoja(ws, cfg, nombre=None):
    """Cuánto se parece esta hoja a un estado de situación financiera.

    Puntúa por lo que la hoja CONTIENE ('Total assets', 'ASSETS', 'Situación
    Financiera'…). El nombre solo suma medio punto: es una pista, no la
    prueba. Así la hoja buena se reconoce aunque la renombren, y una hoja
    que se llame como toca pero no tenga la forma no se cuela.
    """
    marcadores = [_norm(m) for m in cfg["hoja_marcadores"] if _norm(m)]
    texto_hoja = _texto_de_hoja(ws)
    if not texto_hoja:
        return 0.0

    score = float(sum(1 for m in marcadores if m in texto_hoja))
    convencion = _norm(cfg.get("buscar_por_convencion"))
    if convencion and convencion in _norm(nombre or getattr(ws, "title", "")):
        score += 0.5
    try:
        if getattr(ws, "sheet_state", "visible") == "visible":
            score += 0.5
    except Exception:
        pass
    return score


def _elegir_hoja_por_contenido(wb, cfg):
    mejor_nombre, mejor_score = None, 0

    for nombre in wb.sheetnames:
        score = puntuar_hoja(wb[nombre], cfg, nombre)
        if score > mejor_score:
            mejor_nombre, mejor_score = nombre, score

    if mejor_nombre is not None and mejor_score >= UMBRAL_HOJA:
        return mejor_nombre, f"'{mejor_nombre}' elegida por contenido ({mejor_score:g} señales)"

    raise ValueError(
        "No pude identificar la hoja del Estado de Situación Financiera.\n"
        f"No existe una hoja llamada '{cfg['hoja']}' y ninguna otra hoja tiene\n"
        "suficientes marcadores de contenido (p. ej. 'Total assets', 'ASSETS',\n"
        "'Situación Financiera').\n"
        "Indique el nombre exacto en config.json -> \"hoja\"."
    )


# --------------------------------------------------------------------------- #
#  Materialización de la hoja elegida (una sola pasada)
# --------------------------------------------------------------------------- #
def _materializar(ws, cfg):
    # Los limites de config dejan de ser un TOPE y pasan a ser un MINIMO: se
    # barre al menos eso, y ademas todo lo que la hoja diga que ocupa.
    #
    # Antes eran un tope fijo (16 columnas). En cuanto alguien movia la
    # tabla a la derecha —K14, por ejemplo— la columna 'Tipo' caia fuera del
    # barrido y los tipos declarados desaparecian sin un solo aviso: el
    # programa volvia a adivinarlos como si la columna no existiera.
    #
    # Los topes de seguridad evitan que una hoja con basura en la celda
    # XFD1048576 obligue a materializar un millon de filas.
    max_filas = max(int(cfg["max_filas_scan"]), min(ws.max_row or 1, 20000))
    max_cols = max(int(cfg["max_cols_scan"]), min(ws.max_column or 1, 256))
    valores = [[None] * (max_cols + 1)]
    negrita = [[False] * (max_cols + 1)]
    for fila in ws.iter_rows(min_row=1, max_row=max_filas, max_col=max_cols):
        vrow = [None] * (max_cols + 1)
        brow = [False] * (max_cols + 1)
        for c, celda in enumerate(fila, start=1):
            if c > max_cols:
                break
            vrow[c] = celda.value
            try:
                brow[c] = bool(celda.font and celda.font.bold)
            except Exception:
                brow[c] = False
        valores.append(vrow)
        negrita.append(brow)

    n_filas = len(valores) - 1
    while n_filas > 1 and all(v in (None, "") for v in valores[n_filas][1:]):
        n_filas -= 1
    return valores, negrita, n_filas


# --------------------------------------------------------------------------- #
#  Detección de columnas por contenido
# --------------------------------------------------------------------------- #
def detectar_columnas(valores, cfg, n_filas, max_cols):
    forzado = {k: _col_por_letra(v) for k, v in cfg["columnas"].items()}
    col = dict(forzado)

    # 1. rótulos explícitos ('Tipo', 'Note'), estén donde estén
    #
    # Se busca en TODA la altura barrida, no en las 8 primeras filas: si la
    # tabla no arranca arriba del todo, esos rotulos quedaban fuera y la
    # columna 'Tipo' se daba por inexistente. El resultado era mudo y
    # peligroso: todos los tipos volvian a inferirse como si nunca se
    # hubieran declarado. Gana la coincidencia mas alta de la hoja.
    for r in range(1, n_filas + 1):
        for c in range(1, max_cols + 1):
            t = _norm(valores[r][c])
            if not t:
                continue
            if col.get("tipo") is None and t in ("tipo", "type"):
                col["tipo"] = c
            if col.get("nota") is None and t in ("note", "nota", "notes", "notas"):
                col["nota"] = c

    # 2. perfil de cada columna sobre el cuerpo de la hoja
    perfil = {}
    for c in range(1, max_cols + 1):
        n_txt = n_num = 0
        for r in range(2, n_filas + 1):
            v = valores[r][c]
            if _es_numero(v):
                n_num += 1
            elif _texto_no_vacio(v):
                n_txt += 1
        perfil[c] = (n_txt, n_num)

    # 3. columnas de cifras: las dos con más numéricos (empate -> más a la derecha)
    if col.get("actual") is None or col.get("previo") is None:
        con_numeros = [c for c in perfil if perfil[c][1] >= 2]
        con_numeros.sort(key=lambda c: (perfil[c][1], c), reverse=True)
        elegidas = sorted(con_numeros[:2])
        if len(elegidas) >= 2:
            col["actual"] = col.get("actual") or elegidas[0]
            col["previo"] = col.get("previo") or elegidas[1]
        elif len(elegidas) == 1:
            col["actual"] = col.get("actual") or elegidas[0]

    # 4. etiqueta: primera columna con bastante texto, a la izquierda de las cifras
    if col.get("etiqueta") is None:
        limite = col.get("actual") or (max_cols + 1)
        textuales = [c for c in range(1, limite) if perfil[c][0] >= 3]
        col["etiqueta"] = textuales[0] if textuales else 1

    # 5. nota: columna intermedia casi vacía con enteros pequeños
    if col.get("nota") is None and col.get("actual"):
        for c in range(col["etiqueta"] + 1, col["actual"]):
            n_txt, _ = perfil[c]
            enteros_peq = sum(
                1 for r in range(2, n_filas + 1)
                if _es_numero(valores[r][c]) and float(valores[r][c]).is_integer()
                and 0 < valores[r][c] <= 99
            )
            if enteros_peq >= 1 and n_txt <= 2:
                col["nota"] = c
                break

    if not col.get("actual"):
        raise ValueError(
            "No pude identificar las columnas de cifras en la hoja.\n"
            "Defina las letras a mano en config.json -> \"columnas\"\n"
            "(etiqueta / nota / actual / previo / tipo)."
        )

    return {
        "etiqueta": col.get("etiqueta"),
        "nota": col.get("nota"),
        "actual": col.get("actual"),
        "previo": col.get("previo"),
        "tipo": col.get("tipo"),
    }


# --------------------------------------------------------------------------- #
#  Región de datos (fila inicial y final)
# --------------------------------------------------------------------------- #
def detectar_region(valores, cols, cfg, n_filas):
    ce, ca, cp = cols["etiqueta"], cols["actual"], cols["previo"]
    pf = cfg["primera_fila"]

    if isinstance(pf, int) and pf > 0:
        primera = pf
    else:
        # El bloque de cabecera se localiza por contenido; los datos
        # empiezan justo detrás. Antes esto era «las filas 1 a 4 son
        # cabecera», y una tabla que no empezara arriba del todo metía sus
        # propias fechas y su fila de unidades dentro de la tabla del Word.
        hdr = detectar_encabezado(valores, cols, n_filas)
        arranque = max(1, hdr.get("fin", 0) + 1)
        reconocido = hdr.get("fin", 0) > 0

        primera = None
        for r in range(arranque, n_filas + 1):
            et = valores[r][ce]
            va = valores[r][ca] if ca else None
            vp = valores[r][cp] if cp else None
            if not _texto_no_vacio(et):
                continue
            if _norm(et) in ("", "$"):
                continue
            # Si no se reconoció ningún bloque de cabecera se conserva la
            # regla de siempre, para no cambiarle el resultado a una hoja
            # que hoy funciona.
            if reconocido or r > 4 or _es_numero(va) or _es_numero(vp):
                primera = r
                break
        if primera is None:
            primera = arranque if reconocido else 5

    ultima = primera
    for r in range(primera, n_filas + 1):
        if (_texto_no_vacio(valores[r][ce])
                or (ca and _es_numero(valores[r][ca]))
                or (cp and _es_numero(valores[r][cp]))):
            ultima = r
    return primera, ultima


# --------------------------------------------------------------------------- #
#  Encabezado (título, fechas, unidad, moneda)
# --------------------------------------------------------------------------- #
#: Nombres de mes en los dos idiomas del libro. Sirven para reconocer una
#: fila de fechas ("June 5,", "31 de diciembre") sin saber donde esta.
_MESES = (
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
    "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
    "agosto", "septiembre", "setiembre", "octubre", "noviembre", "diciembre",
)


def _es_fecha_cabecera(v):
    from datetime import date, datetime
    if isinstance(v, (datetime, date)):
        return True
    if not isinstance(v, str):
        return False
    s = _norm(v)
    if not s:
        return False
    if any(m in s for m in _MESES):
        return True
    if re.search(r"\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}", s):
        return True
    return bool(re.search(r"\b(19|20)\d{2}\b", s))


def _es_escala_cabecera(v):
    """La fila de la unidad: 1000, 1000000, 'in thousands'…"""
    if _es_numero(v):
        return float(v) in (100.0, 1000.0, 10000.0, 1000000.0)
    if isinstance(v, str):
        return any(p in _norm(v) for p in
                   ("thousand", "million", "miles", "millones", "millon"))
    return False


def _es_estado_cabecera(v):
    """(Unaudited) / (Audited) / (No auditado)…"""
    return isinstance(v, str) and "audit" in _norm(v)


def _es_moneda_cabecera(v):
    if not isinstance(v, str):
        return False
    s = v.strip().strip("()").strip()
    if not s or len(s) > 6:
        return False
    return bool(re.fullmatch(r"[$€£¥]|[A-Za-z]{0,3}\$|USD|COP|CAD|EUR|MXN|GBP",
                             s, re.IGNORECASE))


def detectar_encabezado(valores, cols, n_filas):
    """En qué fila está cada dato de la cabecera. Por CONTENIDO, no por sitio.

    Antes se leían las filas 1, 2, 3 y 4 de la hoja, en ese orden fijo. En
    cuanto alguien movía la tabla —a K14, por ejemplo— esas cuatro celdas
    quedaban vacías y el documento salía con «Al () — comparado con ()» y
    «Cifras expresadas en , en unidades de».

    Devuelve {'fechas': fila, 'escala': fila, 'estado': fila, 'moneda': fila,
    'fin': última fila del bloque}. 'fin' es 0 si no se reconoció ninguna:
    entonces el llamante se queda con la heurística de siempre.
    """
    ce, ca, cp = cols["etiqueta"], cols["actual"], cols["previo"]
    hallado, fin = {}, 0

    for r in range(1, n_filas + 1):
        va = valores[r][ca] if ca else None
        vp = valores[r][cp] if cp else None
        etq = valores[r][ce] if ce else None

        clase = None
        if _es_fecha_cabecera(va) or _es_fecha_cabecera(vp):
            clase = "fechas"
        elif _es_escala_cabecera(va) and not _texto_no_vacio(etq):
            # sin etiqueta: si no, un importe de 1.000 en una linea de
            # detalle se confundiria con la fila de la unidad
            clase = "escala"
        elif _es_estado_cabecera(va) or _es_estado_cabecera(vp):
            clase = "estado"
        elif _es_moneda_cabecera(va) or _es_moneda_cabecera(vp):
            clase = "moneda"

        if clase:
            hallado.setdefault(clase, r)
            fin = max(fin, r)
            continue

        # Una fila con etiqueta Y cifras de verdad ya es un dato: el bloque
        # de cabecera se acabó. Las filas de seccion (texto sin cifras) no
        # cortan: pueden ir intercaladas antes del primer importe.
        if _texto_no_vacio(etq) and (_es_numero(va) or _es_numero(vp)):
            break

    hallado["fin"] = fin
    return hallado


def leer_encabezado(valores, cols, cfg, primera, hdr=None):
    ce, ca, cp = cols["etiqueta"], cols["actual"], cols["previo"]
    hdr = hdr or {}

    def celda(clase, col):
        r = hdr.get(clase)
        if r and col and 0 < r < len(valores):
            return valores[r][col]
        return None

    return {
        "empresa": cfg["empresa"],
        "titulo": _primer_texto(valores, ce, range(1, max(primera, 2))) or "",
        "fecha_actual": celda("fechas", ca) or "",
        "fecha_previa": celda("fechas", cp) or "",
        "miles": celda("escala", ca) or "",
        "estado_actual": texto(celda("estado", ca)),
        "estado_previo": texto(celda("estado", cp)),
        "moneda": celda("moneda", ca) or "",
    }


# --------------------------------------------------------------------------- #
#  Inferencia del tipo de fila
# --------------------------------------------------------------------------- #
def inferir_tipo(vals, bold, cols, cfg):
    ce, cn, ca, cp = cols["etiqueta"], cols["nota"], cols["actual"], cols["previo"]
    et = vals[ce]
    nota = vals[cn] if cn else None
    va = vals[ca] if ca else None
    vp = vals[cp] if cp else None

    et_txt = et.strip() if isinstance(et, str) else ("" if et is None else str(et))
    et_norm = _norm(et_txt)
    nota_norm = _norm(nota)
    hay_valor = (_es_numero(va) or _es_numero(vp)
                 or _marcador_cero(va) or _marcador_cero(vp))
    tiene_nota = _es_numero(nota) or _texto_no_vacio(str(nota) if nota is not None else "")
    en_negrita = bool((ca and bold[ca]) or (cp and bold[cp]) or (ce and bold[ce]))

    # exclusión: el marcador puede estar en la etiqueta o en la columna de nota
    for m in cfg["marcadores_excluir"]:
        mm = _norm(m)
        if mm and (mm in et_norm or mm in nota_norm):
            return "X", "fila de control/cuadre"

    if not et_txt and not hay_valor and not tiene_nota:
        return None, "fila vacía"

    if et_norm.startswith("total"):
        return "T", "empieza por 'Total'"

    if not et_txt and hay_valor:
        return "S", "cifras sin etiqueta (subtotal)"

    if et_txt and not hay_valor:
        if et_txt.isupper() or et_txt.endswith(":") or en_negrita or len(et_txt) <= 40:
            return "H", "etiqueta sin cifras (encabezado)"
        return "N", "texto largo sin cifras (nota)"

    if et_txt and (hay_valor or tiene_nota):
        return "I", "etiqueta con cifras o nota"

    return None, "sin señales suficientes"


# --------------------------------------------------------------------------- #
#  Construcción de las líneas + registro de revisión
# --------------------------------------------------------------------------- #
def construir_lineas(valores, negrita, cols, cfg, primera, ultima, hay_col_tipo):
    ce, cn, ca, cp, ct = (cols["etiqueta"], cols["nota"], cols["actual"],
                          cols["previo"], cols["tipo"])
    lineas, revision = [], []
    sin_tipo, tipo_invalido = [], []
    n_declarados = n_inferidos = 0

    for r in range(primera, ultima + 1):
        vals, bold = valores[r], negrita[r]
        crudo = vals[ct] if ct else None
        declarado = str(crudo).strip().upper() if crudo not in (None, "") else ""

        et_prev = vals[ce] if _texto_no_vacio(vals[ce]) else ""

        if declarado == "X":
            revision.append((r, et_prev, "", "", "", "X", "declarado", "excluido a propósito"))
            continue

        if declarado in TIPOS_VALIDOS:
            tipo, origen, senal = declarado, "declarado", ""
        elif declarado:
            tipo_invalido.append((r, crudo))
            inf, senal = inferir_tipo(vals, bold, cols, cfg)
            if inf in (None, "X"):
                revision.append((r, et_prev, "", "", "", declarado, "tipo inválido",
                                 f"'{crudo}' no es H/I/S/T/N; {senal}"))
                continue
            tipo, origen = inf, "inferido (Tipo inválido)"
        else:
            inf, senal = inferir_tipo(vals, bold, cols, cfg)
            if inf is None:
                continue
            if inf == "X":
                revision.append((r, et_prev, "", "", "", "", "inferido", senal))
                continue
            tipo, origen = inf, "inferido"
            if hay_col_tipo:
                sin_tipo.append(r)

        etiqueta = vals[ce] or ""
        nota = texto(vals[cn]) if cn else ""
        actual = num(vals[ca]) if ca else ""
        previo = num(vals[cp]) if cp else ""
        # *_raw: el valor sin formatear. Lo usa fs_contrato para calcular
        # variaciones (var_abs / var_pct). docxtpl ignora las claves de más.
        lineas.append({"tipo": tipo, "etiqueta": etiqueta, "nota": nota,
                       "actual": actual, "previo": previo,
                       "fila": r,
                       "actual_raw": vals[ca] if ca else None,
                       "previo_raw": vals[cp] if cp else None})
        revision.append((r, etiqueta, nota, actual, previo, tipo, origen, senal))
        if origen == "declarado":
            n_declarados += 1
        else:
            n_inferidos += 1

    return lineas, revision, sin_tipo, tipo_invalido, n_declarados, n_inferidos


def escribir_revision(revision):
    SALIDAS.mkdir(exist_ok=True)
    ruta = SALIDAS / "revisar_tipos.csv"
    with ruta.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["fila", "etiqueta", "nota", "actual", "previo",
                    "tipo", "origen", "señal"])
        for row in revision:
            w.writerow(row)
    return ruta


# --------------------------------------------------------------------------- #
#  Rangos con nombre — la identidad estable de cada línea
# --------------------------------------------------------------------------- #
def _decimales_de(formato):
    """Cuántos decimales pide un formato numérico de Excel ('0.00' -> 2)."""
    m = re.search(r"[0#]\.([0#]+)", str(formato or ""))
    return len(m.group(1)) if m else None


def formatear_valor(valor, formato=None):
    """Da formato a una celda suelta RESPETANDO el formato del propio Excel.

    Es distinto de num(): num() es para las cifras de la tabla del estado,
    siempre en miles y sin decimales. Aquí entran ratios, tipos de cambio,
    porcentajes y fechas, donde los decimales importan. Si el usuario puso
    '0.00' en Excel, aquí salen dos decimales; si puso '0.0%', sale con el
    signo de porcentaje.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sí" if valor else "No"
    if isinstance(valor, str):
        return valor.strip()
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d")
    if not isinstance(valor, (int, float)):
        return str(valor)

    fmt = str(formato or "")
    dec = _decimales_de(fmt)
    if "%" in fmt:
        return f"{valor * 100:,.{dec if dec is not None else 1}f}%"
    if dec is None:
        dec = 0 if float(valor).is_integer() else 2
    if valor < 0:
        return f"({abs(valor):,.{dec}f})"
    return f"{valor:,.{dec}f}"


def _leer_celda(wb, hoja, fila, columna):
    """Lee UNA celda de cualquier hoja. En modo read_only no hay acceso
    aleatorio con ws.cell(), así que se pide esa fila con iter_rows."""
    try:
        ws = wb[hoja]
    except KeyError:
        return None, None
    try:
        for fila_celdas in ws.iter_rows(min_row=fila, max_row=fila,
                                        min_col=columna, max_col=columna):
            for celda in fila_celdas:
                return celda.value, getattr(celda, "number_format", None)
    except Exception:
        return None, None
    return None, None


def leer_rangos_con_nombre(wb, hoja_titulo, cfg):
    """Recorre los nombres 'fs_*' del libro ENTERO.

    Devuelve (por_fila, escalares):
      por_fila   {fila: clave}  para los nombres que caen en la hoja del
                 estado; sirven para dar identidad estable a esas líneas.
      escalares  {clave: (texto_formateado, valor_crudo)}  para todo lo
                 demás: una celda de otra hoja, un ratio, un tipo de
                 cambio, una fecha de corte. Es la vía para llevar al Word
                 cifras que NO son filas de la tabla.

    Un nombre de Excel sobrevive a que se renombre la etiqueta de la fila y
    a que se inserten filas encima: Excel reajusta la referencia solo.

    Solo LEE. Nunca reguardamos el libro con openpyxl: descartaría el valor
    cacheado de todas las fórmulas. Para crear nombres, use la orden
    'nombrar' de fs_documento.py, que lo hace con Excel.
    """
    prefijo = str(cfg.get("prefijo_rangos") or "").strip()
    if not prefijo:
        return {}, {}

    por_fila, escalares = {}, {}
    try:
        items = list(wb.defined_names.items())
    except AttributeError:                       # openpyxl < 3.1
        items = [(dn.name, dn) for dn in wb.defined_names.definedName]

    for nombre, dn in items:
        if not nombre.lower().startswith(prefijo.lower()):
            continue
        clave = nombre[len(prefijo):].strip().lower()
        if not clave:
            continue
        try:
            destinos = list(dn.destinations)
        except Exception:
            continue                              # #REF! y otros nombres rotos

        for hoja, coord in destinos:
            try:
                min_col, min_fila, _, _ = range_boundaries(coord.replace("$", ""))
            except Exception:
                continue
            if min_fila is None or min_col is None:
                continue
            if hoja == hoja_titulo:
                # Puede ser una línea del estado; se decide más tarde, al
                # saber qué filas entraron en la región de datos.
                por_fila.setdefault(min_fila, clave)
            valor, formato = _leer_celda(wb, hoja, min_fila, min_col)
            escalares[clave] = (formatear_valor(valor, formato), valor)
            break
    return por_fila, escalares
# --------------------------------------------------------------------------- #
#  Orquestador de lectura
# --------------------------------------------------------------------------- #
def leer_contexto(ruta_xlsx, cfg):
    wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
    try:
        # El nombre de config.json es una PISTA que se comprueba, no una
        # orden. Se mira primero la hoja que se llama así, pero solo se
        # acepta si además tiene forma de estado de situación financiera. Si
        # no la tiene —o si no existe— se elige por contenido. Así el libro
        # sigue funcionando aunque renombren la hoja, y una hoja que se
        # llame 'FS' sin serlo no arrastra al resto.
        objetivo = cfg["hoja"]
        nombre = como_hoja = None
        if objetivo and objetivo in wb.sheetnames:
            puntos = puntuar_hoja(wb[objetivo], cfg, objetivo)
            if puntos >= UMBRAL_HOJA:
                nombre = objetivo
                como_hoja = (f"'{objetivo}' por nombre, confirmada por su "
                             f"contenido ({puntos:g} señales)")
        if nombre is None:
            nombre, como_hoja = _elegir_hoja_por_contenido(wb, cfg)

        wb.close()
        wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
        ws = wb[nombre]

        hoja_titulo = ws.title
        valores, negrita, n_filas = _materializar(ws, cfg)
        # El ancho REAL de lo materializado, no el de config: _materializar
        # amplia el barrido hasta donde llegue la hoja, y si aqui se volviera
        # a pasar cfg["max_cols_scan"] la deteccion se quedaria ciega
        # justamente en las columnas recien ganadas.
        ancho = len(valores[0]) - 1
        cols = detectar_columnas(valores, cfg, n_filas, ancho)
        primera, ultima = detectar_region(valores, cols, cfg, n_filas)
        # La misma lectura por contenido que usa detectar_region, para que
        # las fechas y la moneda salgan de donde esten de verdad.
        hdr = detectar_encabezado(valores, cols, n_filas)
        ctx = leer_encabezado(valores, cols, cfg, primera, hdr)
        hay_col_tipo = cols["tipo"] is not None
        (lineas, revision, sin_tipo, tipo_invalido,
         n_declarados, n_inferidos) = construir_lineas(
            valores, negrita, cols, cfg, primera, ultima, hay_col_tipo)

        # Identidad estable por rango con nombre. La línea que tenga un
        # nombre 'fs_*' apuntando a su fila lleva esa clave; el resto cae
        # en la clave derivada de la etiqueta (frágil ante renombrados).
        rangos_por_fila, escalares_libro = leer_rangos_con_nombre(
            wb, hoja_titulo, cfg)
        n_con_rango = 0
        claves_de_lineas = set()
        for linea in lineas:
            clave_rango = rangos_por_fila.get(linea.get("fila"))
            if clave_rango:
                linea["clave"] = clave_rango
                linea["clave_origen"] = "rango"
                claves_de_lineas.add(clave_rango)
                n_con_rango += 1
            else:
                linea["clave_origen"] = "etiqueta"

        # Lo que no es una fila del estado queda como escalar suelto: una
        # celda de otra hoja, un ratio, un tipo de cambio, una fecha de
        # corte. Se expone con el campo 'actual' para poder intercalarlo en
        # la redacción igual que cualquier cifra de la tabla.
        escalares = {k: v for k, v in escalares_libro.items()
                     if k not in claves_de_lineas}
    finally:
        wb.close()

    if not lineas:
        raise ValueError(
            f"Identifiqué la hoja ({hoja_titulo}, {como_hoja}) y las columnas,\n"
            f"pero no obtuve ninguna línea en la región filas {primera}–{ultima}.\n"
            "Posibles causas: el Excel se guardó sin valores cacheados de las\n"
            "fórmulas, o las columnas detectadas no son las correctas.\n"
            "Revise config.json (columnas / primera_fila)."
        )

    ctx["lineas"] = lineas
    ctx["escalares"] = escalares
    ctx["_meta"] = {
        "hoja": hoja_titulo,
        "como_hoja": como_hoja,
        "columnas": {k: (get_column_letter(v) if v else "—") for k, v in cols.items()},
        "region": (primera, ultima),
        "hay_col_tipo": hay_col_tipo,
        "n_declarados": n_declarados,
        "n_inferidos": n_inferidos,
        "n_con_rango": n_con_rango,
        "n_escalares": len(escalares),
    }
    ctx["_avisos"] = {
        "sin_tipo": sin_tipo,
        "tipo_invalido": tipo_invalido,
        "revision": revision,
    }
    return ctx


def encontrar_excel_por_convencion(cfg):
    clave = cfg["buscar_por_convencion"]
    # La carpeta del .exe / raíz del proyecto primero; ejemplos\ después,
    # para que el libro de muestra no gane a uno que el usuario haya dejado.
    candidatos = []
    for carpeta in (BASE, BASE / "ejemplos"):
        if not carpeta.is_dir():
            continue
        candidatos += [
            p for p in carpeta.glob("*.xlsx")
            if clave.lower() in p.stem.lower() and not p.name.startswith("~$")
        ]
    candidatos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidatos:
        icono = "GeneradorFS.exe" if getattr(sys, "frozen", False) else "generar.bat"
        raise ValueError(
            "No encontré ningún Excel en esta carpeta.\n\n"
            f"Haga una de estas dos cosas:\n"
            f"  1) Arrastre su archivo de Excel encima de {icono}, o\n"
            f"  2) Copie el Excel a esta misma carpeta (su nombre debe contener "
            f"'{clave}') y vuelva a ejecutar."
        )
    return candidatos[0]


# --------------------------------------------------------------------------- #
#  Ejecución
# --------------------------------------------------------------------------- #
def ejecutar(argv):
    cfg = cargar_config()
    args = [a for a in argv[1:] if not a.startswith("--")]
    flags = {a.lower() for a in argv[1:] if a.startswith("--")}
    solo_revisar = "--revisar" in flags

    if args:
        xlsx = Path(args[0]).resolve()
    else:
        xlsx = encontrar_excel_por_convencion(cfg)
        print(f"(Sin archivo indicado: usando '{xlsx.name}' por convención de nombre)")

    if len(args) >= 2:
        plantilla = Path(args[1]).resolve()
    else:
        # junto al .exe / en la raíz primero (permite sustituirla sin
        # recompilar), luego plantillas\ y por último la embebida.
        plantilla = buscar_recurso(cfg["plantilla"]) or (BASE / cfg["plantilla"])

    if not xlsx.exists():
        raise ValueError(f"No se encontró el libro de Excel:\n  {xlsx}")
    if not solo_revisar and not plantilla.exists():
        raise ValueError(f"No se encontró la plantilla de Word:\n  {plantilla}")

    ctx = leer_contexto(xlsx, cfg)
    meta = ctx.pop("_meta")
    avisos = ctx.pop("_avisos")

    sello = datetime.now().strftime("%Y%m%d-%H%M")
    csv_ruta = escribir_revision(avisos["revision"])

    destino = None
    if not solo_revisar:
        entorno = SandboxedEnvironment(autoescape=True)
        doc = DocxTemplate(plantilla)
        doc.render(ctx, entorno)
        sha = hashlib.sha256(xlsx.read_bytes()).hexdigest()[:12]
        doc.docx.core_properties.comments = (
            f"origen={xlsx.name} sha256={sha} hoja={meta['hoja']} "
            f"plantilla={plantilla.name} generado={sello}"
        )
        SALIDAS.mkdir(exist_ok=True)
        destino = SALIDAS / f"estado_situacion_financiera_{sanear(ctx['fecha_actual'])}_{sello}.docx"
        doc.save(destino)

    cols = meta["columnas"]
    print()
    print("=" * 64)
    print(" REVISIÓN GENERADA" if solo_revisar else " DOCUMENTO GENERADO CORRECTAMENTE")
    print("=" * 64)
    if destino:
        print(f" Archivo:            {destino.name}")
        print(f" Carpeta:            {destino.parent}")
    print(f" Hoja usada:         {meta['hoja']}  ({meta['como_hoja']})")
    print(f" Columnas:           etiqueta={cols['etiqueta']}  nota={cols['nota']}  "
          f"actual={cols['actual']}  previo={cols['previo']}  tipo={cols['tipo']}")
    print(f" Región de datos:    filas {meta['region'][0]}–{meta['region'][1]}")
    print(f" Líneas trasladadas: {len(ctx['lineas'])}   "
          f"({meta['n_declarados']} declaradas, {meta['n_inferidos']} inferidas)")
    print(f" Revisión de tipos:  {csv_ruta}")

    if not meta["hay_col_tipo"]:
        print()
        print(" NOTA — la hoja no tiene columna 'Tipo': TODOS los tipos se")
        print("        infirieron. Abra revisar_tipos.csv y confírmelos antes")
        print("        de usar el documento. Para fijarlos, agregue una columna")
        print("        'Tipo' a la hoja con H/I/S/T/N/X.")

    if avisos["sin_tipo"]:
        print()
        print(" AVISO — filas con datos pero SIN 'Tipo' declarado (se infirió):")
        print("   Fila(s):", ", ".join(str(r) for r in avisos["sin_tipo"]))

    if avisos["tipo_invalido"]:
        print()
        print(" AVISO — filas con un 'Tipo' no reconocido (se infirió en su lugar):")
        for r, t in avisos["tipo_invalido"]:
            print(f"   Fila {r}: '{t}'  (válidos: H, I, S, T, N, X)")

    print("=" * 64)
    # El llamante (fs_menu) lo necesita para ofrecer «Abrir la carpeta»:
    # el archivo se crea en salidas\ JUNTO AL .exe, que no siempre es donde
    # el usuario cree que está.
    return destino


def _pausa():
    """Espera Enter solo si hay una consola interactiva (doble clic / .bat).
    Si el programa se llama desde un script o una tubería, no pausa."""
    try:
        if sys.stdin and sys.stdin.isatty():
            input("\nPresione Enter para cerrar...")
    except (EOFError, OSError):
        pass


def main():
    try:
        ejecutar(sys.argv)
    except ValueError as e:
        print()
        print("=" * 64)
        print(" NO SE PUDO GENERAR EL DOCUMENTO")
        print("=" * 64)
        print(str(e))
        print("=" * 64)
        _pausa()
        sys.exit(1)
    except Exception:
        print()
        print("=" * 64)
        print(" OCURRIÓ UN ERROR INESPERADO — copie este texto para soporte")
        print("=" * 64)
        traceback.print_exc()
        print("=" * 64)
        _pausa()
        sys.exit(1)
    else:
        _pausa()


if __name__ == "__main__":
    main()
